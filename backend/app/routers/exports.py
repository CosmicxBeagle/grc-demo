"""
Export endpoints — returns styled .xlsx workbooks for auditors.
"""
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from datetime import date as _date

from app.db.database import get_db
from app.auth.permissions import require_permission
from app.models.models import User, ControlException, TreatmentPlan, TreatmentMilestone, Risk
from app.services.services import AuditService
from app.repositories.repositories import (
    ControlRepository, DeficiencyRepository,
    RiskRepository, TestCycleRepository, AssignmentRepository,
)

router = APIRouter(
    prefix="/exports",
    tags=["exports"],
    dependencies=[Depends(require_permission("reports:export"))],
)


# ── Style helpers ──────────────────────────────────────────────────────────

BRAND_DARK  = "1E3A5F"   # dark navy
BRAND_MID   = "2563EB"   # brand blue
HEADER_BG   = PatternFill("solid", fgColor=BRAND_DARK)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Calibri", bold=True, color=BRAND_DARK, size=14)
LABEL_FONT  = Font(name="Calibri", bold=True, color="374151", size=10)
BODY_FONT   = Font(name="Calibri", size=10)

ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")   # very light blue
RED_FILL    = PatternFill("solid", fgColor="FEE2E2")
ORANGE_FILL = PatternFill("solid", fgColor="FFEDD5")
YELLOW_FILL = PatternFill("solid", fgColor="FEF9C3")
GREEN_FILL  = PatternFill("solid", fgColor="DCFCE7")

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers)


def _header_row(ws, columns: list[tuple[str, int]], row: int = 1):
    """Write a styled header row. columns = [(label, width), ...]"""
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill   = HEADER_BG
        cell.font   = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _body_row(ws, values: list, row: int, alt: bool, fill_override=None):
    base_fill = ALT_FILL if alt else None
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font   = BODY_FONT
        cell.border = BORDER
        cell.alignment = LEFT
        if fill_override:
            cell.fill = fill_override
        elif base_fill:
            cell.fill = base_fill


def _title_block(ws, title: str, subtitle: str):
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = TITLE_FONT
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(name="Calibri", italic=True, color="6B7280", size=10)


def _severity_fill(severity: str):
    return {"critical": RED_FILL, "high": ORANGE_FILL, "medium": YELLOW_FILL}.get(severity, None)


def _score_fill(score: int):
    if score >= 20: return RED_FILL
    if score >= 15: return ORANGE_FILL
    if score >= 9:  return YELLOW_FILL
    return GREEN_FILL


# ── Controls export ────────────────────────────────────────────────────────

@router.get("/controls")
def export_controls(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_permission("reports:export"))):
    controls = ControlRepository(db).get_all()
    wb = Workbook()

    # Sheet 1 — Control Library
    ws = wb.active
    ws.title = "Control Library"
    ws.freeze_panes = "A5"

    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _title_block(ws, "Control Library", f"Exported {generated}  |  {len(controls)} controls")

    cols = [
        ("Control ID", 14), ("Title", 42), ("Type", 14),
        ("Frequency", 14), ("Owner", 20), ("Status", 12),
        ("Frameworks", 30), ("Created", 16),
    ]
    _header_row(ws, cols, row=4)

    for i, c in enumerate(controls):
        frameworks = ", ".join(sorted({m.framework for m in c.mappings}))
        _body_row(ws, [
            c.control_id, c.title,
            (c.control_type or "").capitalize(),
            (c.frequency or "").capitalize(),
            c.owner or "",
            c.status.capitalize(),
            frameworks,
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
        ], row=5 + i, alt=bool(i % 2))

    # Sheet 2 — Framework Mappings
    ws2 = wb.create_sheet("Framework Mappings")
    ws2.freeze_panes = "A4"
    _title_block(ws2, "Framework Mappings", f"All control-to-framework cross-references  |  Exported {generated}")

    cols2 = [
        ("Control ID", 14), ("Control Title", 38), ("Framework", 12),
        ("Version", 12), ("Reference", 18), ("Description", 55),
    ]
    _header_row(ws2, cols2, row=3)

    row = 4
    alt = False
    for c in controls:
        for m in sorted(c.mappings, key=lambda x: (x.framework, x.framework_ref)):
            _body_row(ws2, [
                c.control_id, c.title,
                m.framework,
                m.framework_version or "",
                m.framework_ref,
                m.framework_description or "",
            ], row=row, alt=alt)
            row += 1
            alt = not alt

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="control_library", request=request)
    return _xlsx_response(wb, f"control_library_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Deficiency Register export ─────────────────────────────────────────────

@router.get("/deficiencies")
def export_deficiencies(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_permission("reports:export"))):
    deficiencies = DeficiencyRepository(db).get_all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Deficiency Register"
    ws.freeze_panes = "A4"

    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _title_block(ws, "Deficiency Register", f"Exported {generated}  |  {len(deficiencies)} deficiencies")

    cols = [
        ("ID", 6), ("Title", 38), ("Severity", 12), ("Status", 18),
        ("Due Date", 14), ("Remediation Plan", 50),
        ("Created", 14), ("Updated", 14),
    ]
    _header_row(ws, cols, row=3)

    STATUS_LABEL = {
        "open": "Open", "in_remediation": "In Remediation",
        "remediated": "Remediated", "risk_accepted": "Risk Accepted",
    }

    for i, d in enumerate(deficiencies):
        sev_fill = _severity_fill(d.severity)
        _body_row(ws, [
            d.id,
            d.title,
            d.severity.capitalize(),
            STATUS_LABEL.get(d.status, d.status),
            str(d.due_date) if d.due_date else "",
            d.remediation_plan or "",
            d.created_at.strftime("%Y-%m-%d") if d.created_at else "",
            d.updated_at.strftime("%Y-%m-%d") if d.updated_at else "",
        ], row=4 + i, alt=bool(i % 2), fill_override=sev_fill)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="deficiency_register", request=request)
    return _xlsx_response(wb, f"deficiency_register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Risk Register export ───────────────────────────────────────────────────

@router.get("/risks")
def export_risks(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_permission("reports:export"))):
    risks = RiskRepository(db).get_all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"
    ws.freeze_panes = "A4"

    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _title_block(ws, "Risk Register", f"Exported {generated}  |  {len(risks)} risks")

    cols = [
        ("ID", 6), ("Risk Name", 36), ("Asset", 22), ("Threat", 24),
        ("Likelihood", 12), ("Impact", 10), ("Score", 10), ("Rating", 12),
        ("Treatment", 14), ("Status", 14), ("Owner", 20),
        ("Mitigating Controls", 40),
    ]
    _header_row(ws, cols, row=3)

    def rating(score):
        if score >= 20: return "Critical"
        if score >= 15: return "High"
        if score >= 9:  return "Medium"
        return "Low"

    for i, r in enumerate(risks):
        score = r.likelihood * r.impact
        controls_str = ", ".join(
            rc.control.control_id for rc in r.controls if rc.control
        )
        score_fill = _score_fill(score)
        _body_row(ws, [
            r.id,
            r.name,
            r.asset.name if r.asset else "",
            r.threat.name if r.threat else "",
            r.likelihood,
            r.impact,
            score,
            rating(score),
            (r.treatment or "").capitalize(),
            r.status.capitalize(),
            r.owner or "",
            controls_str,
        ], row=4 + i, alt=bool(i % 2), fill_override=score_fill)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="risk_register", request=request)
    return _xlsx_response(wb, f"risk_register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Test Cycle Report export ───────────────────────────────────────────────

@router.get("/test-cycles/{cycle_id}")
def export_test_cycle(cycle_id: int, request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_permission("reports:export"))):
    cycle = TestCycleRepository(db).get_by_id(cycle_id)
    if not cycle:
        from fastapi import HTTPException
        raise HTTPException(404, "Test cycle not found")

    wb = Workbook()
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    def kv(label, value, row):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = LABEL_FONT
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = BODY_FONT

    ws.cell(row=1, column=1, value=cycle.name).font = TITLE_FONT
    ws.row_dimensions[1].height = 28
    ws.cell(row=2, column=1, value=f"Test Cycle Report  |  Exported {generated}").font = Font(
        name="Calibri", italic=True, color="6B7280", size=10)

    kv("Cycle Name",   cycle.name,                                 row=4)
    kv("Status",       cycle.status.capitalize(),                  row=5)
    kv("Brand",        cycle.brand or "—",                        row=6)
    kv("Start Date",   str(cycle.start_date) if cycle.start_date else "—", row=7)
    kv("End Date",     str(cycle.end_date)   if cycle.end_date   else "—", row=8)
    kv("Description",  cycle.description or "—",                   row=9)

    assignments = cycle.assignments or []
    total   = len(assignments)
    complete= sum(1 for a in assignments if a.status == "complete")
    failed  = sum(1 for a in assignments if a.status == "failed")
    in_prog = sum(1 for a in assignments if a.status == "in_progress")
    review  = sum(1 for a in assignments if a.status == "needs_review")
    not_st  = sum(1 for a in assignments if a.status == "not_started")

    kv("Total Controls",  total,    row=11)
    kv("Complete",        complete, row=12)
    kv("Needs Review",    review,   row=13)
    kv("In Progress",     in_prog,  row=14)
    kv("Not Started",     not_st,   row=15)
    kv("Failed",          failed,   row=16)

    completion_pct = f"{round(complete / total * 100)}%" if total else "N/A"
    kv("Completion",  completion_pct, row=17)

    # ── Sheet 2: Assignments ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Assignments")
    ws2.freeze_panes = "A3"

    _title_block(ws2, f"Assignments — {cycle.name}", f"Exported {generated}")

    cols2 = [
        ("Control ID", 14), ("Control Title", 38), ("Type", 14),
        ("Tester", 20), ("Reviewer", 20), ("Status", 16),
        ("Evidence", 10), ("Deficiencies", 13), ("Tester Notes", 40), ("Reviewer Comments", 40),
    ]
    _header_row(ws2, cols2, row=3)

    STATUS_FILL = {
        "complete":     GREEN_FILL,
        "failed":       RED_FILL,
        "needs_review": YELLOW_FILL,
        "in_progress":  PatternFill("solid", fgColor="DBEAFE"),
    }

    for i, a in enumerate(sorted(assignments, key=lambda x: (x.control.control_id if x.control else ""))):
        fill = STATUS_FILL.get(a.status)
        _body_row(ws2, [
            a.control.control_id if a.control else "",
            a.control.title if a.control else "",
            (a.control.control_type or "").capitalize() if a.control else "",
            a.tester.display_name if a.tester else "",
            a.reviewer.display_name if a.reviewer else "",
            a.status.replace("_", " ").title(),
            len(a.evidence),
            len(a.deficiencies),
            a.tester_notes or "",
            a.reviewer_comments or "",
        ], row=4 + i, alt=bool(i % 2), fill_override=fill)

    # ── Sheet 3: Deficiencies ─────────────────────────────────────────────
    all_deficiencies = [d for a in assignments for d in (a.deficiencies or [])]

    ws3 = wb.create_sheet("Deficiencies")
    ws3.freeze_panes = "A4"
    _title_block(ws3, f"Deficiencies — {cycle.name}",
                 f"Exported {generated}  |  {len(all_deficiencies)} deficiencies")

    cols3 = [
        ("Control ID", 14), ("Control Title", 32), ("Deficiency Title", 36),
        ("Severity", 12), ("Status", 18), ("Due Date", 14), ("Remediation Plan", 50),
    ]
    _header_row(ws3, cols3, row=3)

    row = 4
    for a in sorted(assignments, key=lambda x: (x.control.control_id if x.control else "")):
        for d in (a.deficiencies or []):
            sev_fill = _severity_fill(d.severity)
            _body_row(ws3, [
                a.control.control_id if a.control else "",
                a.control.title if a.control else "",
                d.title,
                d.severity.capitalize(),
                d.status.replace("_", " ").title(),
                str(d.due_date) if d.due_date else "",
                d.remediation_plan or "",
            ], row=row, alt=bool((row - 4) % 2), fill_override=sev_fill)
            row += 1

    if row == 4:
        ws3.cell(row=4, column=1, value="No deficiencies recorded for this cycle.").font = Font(
            name="Calibri", italic=True, color="9CA3AF", size=10)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name=f"test_cycle_{cycle_id}", request=request)
    filename = f"test_cycle_{cycle_id}_{cycle.name.replace(' ', '_')[:30]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(wb, filename)


# ── SOX ITGC Scoping export ────────────────────────────────────────────────

ITGC_DOMAINS = [
    "Access Controls",
    "Change Management",
    "Computer Operations",
    "Program Development",
]

ALL_ASSERTIONS = [
    "Completeness",
    "Accuracy",
    "Existence",
    "Authorization",
    "Valuation",
    "Presentation & Disclosure",
]

DOMAIN_FILL = {
    "Access Controls":     PatternFill("solid", fgColor="DBEAFE"),  # blue-100
    "Change Management":   PatternFill("solid", fgColor="EDE9FE"),  # purple-100
    "Computer Operations": PatternFill("solid", fgColor="FFEDD5"),  # orange-100
    "Program Development": PatternFill("solid", fgColor="DCFCE7"),  # green-100
}


@router.get("/sox")
def export_sox(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_permission("reports:export"))):
    controls = ControlRepository(db).get_all()
    in_scope = [c for c in controls if c.sox_in_scope]
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()

    # ── Sheet 1: SOX Control Matrix ───────────────────────────────────────
    ws = wb.active
    ws.title = "SOX Control Matrix"
    ws.freeze_panes = "A5"

    _title_block(ws, "SOX ITGC Scoping — Control Matrix",
                 f"Exported {generated}  |  {len(in_scope)} controls in scope")

    # Assertion columns header (C A E Au V P)
    assertion_abbrev = ["C", "A", "E", "Au", "V", "P&D"]
    cols = [
        ("Control ID", 14), ("Title", 38), ("ITGC Domain", 22),
        ("Systems Covered", 30), ("Owner", 20), ("Type", 14), ("Frequency", 14),
    ] + [(a, 6) for a in assertion_abbrev] + [("Frameworks", 20)]
    _header_row(ws, cols, row=4)

    # Add assertion legend row just above header
    legend_cell = ws.cell(row=3, column=1, value="Assertion columns: C=Completeness  A=Accuracy  E=Existence  Au=Authorization  V=Valuation  P&D=Presentation & Disclosure")
    legend_cell.font = Font(name="Calibri", italic=True, color="6B7280", size=9)

    for i, c in enumerate(sorted(in_scope, key=lambda x: (x.sox_itgc_domain or "ZZZ", x.control_id))):
        active = set(a.strip() for a in (c.sox_assertions or "").split(",") if a.strip())
        frameworks = ", ".join(sorted({m.framework for m in c.mappings}))
        domain_fill = DOMAIN_FILL.get(c.sox_itgc_domain or "", ALT_FILL if i % 2 else None)

        assertion_values = ["Y" if a in active else "" for a in ALL_ASSERTIONS]

        row_data = [
            c.control_id, c.title,
            c.sox_itgc_domain or "",
            c.sox_systems or "",
            c.owner or "",
            (c.control_type or "").capitalize(),
            (c.frequency or "").capitalize(),
        ] + assertion_values + [frameworks]

        _body_row(ws, row_data, row=5 + i, alt=bool(i % 2), fill_override=domain_fill)

        # Center the assertion columns (cols 8-13)
        for col_offset in range(len(ALL_ASSERTIONS)):
            cell = ws.cell(row=5 + i, column=8 + col_offset)
            cell.alignment = CENTER
            if cell.value == "Y":
                cell.font = Font(name="Calibri", bold=True, color="1D4ED8", size=10)

    # ── Sheet 2: Domain Summary ───────────────────────────────────────────
    ws2 = wb.create_sheet("Domain Summary")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 40

    _title_block(ws2, "SOX ITGC — Domain Summary", f"Exported {generated}")

    header_cols = [("ITGC Domain", 26), ("Control Count", 14), ("Systems", 40)]
    _header_row(ws2, header_cols, row=3)

    for i, domain in enumerate(ITGC_DOMAINS):
        domain_controls = [c for c in in_scope if c.sox_itgc_domain == domain]
        systems = sorted({
            s.strip()
            for c in domain_controls
            for s in (c.sox_systems or "").split(",")
            if s.strip()
        })
        fill = DOMAIN_FILL.get(domain, ALT_FILL)
        _body_row(ws2, [domain, len(domain_controls), ", ".join(systems)],
                  row=4 + i, alt=False, fill_override=fill)

    # ── Sheet 3: Assertion Coverage ───────────────────────────────────────
    ws3 = wb.create_sheet("Assertion Coverage")
    ws3.freeze_panes = "B2"

    # Build assertion × domain matrix
    ws3.cell(row=1, column=1, value="Assertion \\ Domain").font = LABEL_FONT
    for ci, domain in enumerate(ITGC_DOMAINS, start=2):
        cell = ws3.cell(row=1, column=ci, value=domain)
        cell.fill = DOMAIN_FILL.get(domain, HEADER_BG)
        cell.font = Font(name="Calibri", bold=True, color="111827", size=10)
        cell.alignment = CENTER
        ws3.column_dimensions[get_column_letter(ci)].width = 22

    ws3.column_dimensions["A"].width = 26

    for ri, assertion in enumerate(ALL_ASSERTIONS, start=2):
        ws3.cell(row=ri, column=1, value=assertion).font = LABEL_FONT
        for ci, domain in enumerate(ITGC_DOMAINS, start=2):
            count = sum(
                1 for c in in_scope
                if c.sox_itgc_domain == domain
                and assertion in {a.strip() for a in (c.sox_assertions or "").split(",")}
            )
            cell = ws3.cell(row=ri, column=ci, value=count if count else "")
            cell.alignment = CENTER
            cell.font = BODY_FONT
            cell.border = BORDER
            if count:
                cell.fill = DOMAIN_FILL.get(domain, ALT_FILL)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="sox_itgc_scoping", request=request)
    return _xlsx_response(wb, f"sox_itgc_scoping_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Exceptions Register export ─────────────────────────────────────────────

EXCEPTION_STATUS_FILL = {
    "approved":         GREEN_FILL,
    "rejected":         RED_FILL,
    "pending_approval": YELLOW_FILL,
    "expired":          PatternFill("solid", fgColor="E5E7EB"),  # gray
}

@router.get("/exceptions")
def export_exceptions(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports:export")),
):
    exceptions = db.query(ControlException).order_by(ControlException.created_at.desc()).all()
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()
    ws = wb.active
    ws.title = "Exceptions Register"
    ws.freeze_panes = "A5"

    _title_block(ws, "Exceptions Register", f"Exported {generated}  |  {len(exceptions)} exceptions")

    cols = [
        ("ID", 6), ("Control", 14), ("Title", 36), ("Type", 20),
        ("Risk Level", 12), ("Status", 18),
        ("Requested By", 22), ("Approved By", 22), ("Approver Notes", 40),
        ("Expiry Date", 14), ("Created", 14),
    ]
    _header_row(ws, cols, row=4)

    STATUS_LABEL = {
        "draft": "Draft", "pending_approval": "Pending Approval",
        "approved": "Approved", "rejected": "Rejected", "expired": "Expired",
    }

    for i, e in enumerate(exceptions):
        fill = EXCEPTION_STATUS_FILL.get(e.status)
        _body_row(ws, [
            e.id,
            e.control.control_id if e.control else "",
            e.title,
            (e.exception_type or "").replace("_", " ").title(),
            (e.risk_level or "").capitalize(),
            STATUS_LABEL.get(e.status, e.status),
            e.requester.display_name if e.requester else "",
            e.approver.display_name if e.approver else "",
            e.approver_notes or "",
            str(e.expiry_date) if e.expiry_date else "",
            e.created_at.strftime("%Y-%m-%d") if e.created_at else "",
        ], row=5 + i, alt=bool(i % 2), fill_override=fill)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="exceptions_register", request=request)
    return _xlsx_response(wb, f"exceptions_register_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Treatment Plans export ─────────────────────────────────────────────────

PLAN_STATUS_FILL = {
    "completed":   GREEN_FILL,
    "in_progress": PatternFill("solid", fgColor="DBEAFE"),
    "on_hold":     YELLOW_FILL,
    "cancelled":   PatternFill("solid", fgColor="E5E7EB"),
}

MILESTONE_STATUS_FILL = {
    "completed": GREEN_FILL,
    "overdue":   RED_FILL,
    "in_progress": PatternFill("solid", fgColor="DBEAFE"),
}

@router.get("/treatment-plans")
def export_treatment_plans(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports:export")),
):
    plans = (
        db.query(TreatmentPlan)
        .order_by(TreatmentPlan.created_at.desc())
        .all()
    )
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()

    # Sheet 1 — Treatment Plans
    ws = wb.active
    ws.title = "Treatment Plans"
    ws.freeze_panes = "A4"
    _title_block(ws, "Treatment Plans", f"Exported {generated}  |  {len(plans)} plans")

    cols = [
        ("Risk", 36), ("Strategy", 14), ("Status", 16),
        ("Owner", 22), ("Target Date", 14), ("Description", 50),
        ("Milestones", 10), ("Complete", 10),
    ]
    _header_row(ws, cols, row=3)

    for i, p in enumerate(plans):
        total_ms = len(p.milestones)
        done_ms  = sum(1 for m in p.milestones if m.status == "completed")
        fill = PLAN_STATUS_FILL.get(p.status)
        _body_row(ws, [
            p.risk.name if p.risk else "",
            (p.strategy or "").capitalize(),
            (p.status or "").replace("_", " ").title(),
            p.owner.display_name if p.owner else "",
            str(p.target_date) if p.target_date else "",
            p.description or "",
            total_ms,
            done_ms,
        ], row=4 + i, alt=bool(i % 2), fill_override=fill)

    # Sheet 2 — Milestones
    ws2 = wb.create_sheet("Milestones")
    ws2.freeze_panes = "A4"
    _title_block(ws2, "Treatment Plan Milestones", f"Exported {generated}")

    cols2 = [
        ("Risk", 30), ("Milestone", 38), ("Assigned To", 22),
        ("Due Date", 14), ("Status", 16),
    ]
    _header_row(ws2, cols2, row=3)

    row = 4
    for p in plans:
        for m in p.milestones:
            fill = MILESTONE_STATUS_FILL.get(m.status)
            _body_row(ws2, [
                p.risk.name if p.risk else "",
                m.title,
                m.assigned_to.display_name if m.assigned_to else "",
                str(m.due_date) if m.due_date else "",
                (m.status or "").replace("_", " ").title(),
            ], row=row, alt=bool((row - 4) % 2), fill_override=fill)
            row += 1

    if row == 4:
        ws2.cell(row=4, column=1, value="No milestones recorded.").font = Font(
            name="Calibri", italic=True, color="9CA3AF", size=10)

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="treatment_plans", request=request)
    return _xlsx_response(wb, f"treatment_plans_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


# ── Risk Aging export ──────────────────────────────────────────────────────

def _age_bucket(days: int) -> str:
    if days <= 30:   return "0–30 days"
    if days <= 60:   return "31–60 days"
    if days <= 90:   return "61–90 days"
    if days <= 180:  return "91–180 days"
    if days <= 365:  return "181–365 days"
    return "365+ days"

AGE_FILLS = [
    GREEN_FILL,
    PatternFill("solid", fgColor="D1FAE5"),   # light green  31-60
    YELLOW_FILL,                               # yellow       61-90
    ORANGE_FILL,                               # orange       91-180
    RED_FILL,                                  # red          181-365
    PatternFill("solid", fgColor="7F1D1D"),   # dark red     365+  (white text below)
]
AGE_BUCKET_LABELS = ["0–30 days", "31–60 days", "61–90 days", "91–180 days", "181–365 days", "365+ days"]

def _age_fill(days: int):
    if days <= 30:   return AGE_FILLS[0]
    if days <= 60:   return AGE_FILLS[1]
    if days <= 90:   return AGE_FILLS[2]
    if days <= 180:  return AGE_FILLS[3]
    if days <= 365:  return AGE_FILLS[4]
    return AGE_FILLS[5]


@router.get("/risk-aging")
def export_risk_aging(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports:export")),
):
    risks = db.query(Risk).order_by(Risk.created_at).all()
    today = _date.today()
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()

    # Sheet 1 — Risk Aging Detail
    ws = wb.active
    ws.title = "Risk Aging"
    ws.freeze_panes = "A5"
    _title_block(ws, "Risk Aging Report",
                 f"Exported {generated}  |  {len(risks)} risks  |  As-of {today}")

    cols = [
        ("ID", 6), ("Risk Name", 36), ("Asset", 22), ("Threat", 24),
        ("Inherent Score", 14), ("Residual Score", 14), ("Rating", 12),
        ("Treatment", 14), ("Status", 14), ("Owner", 20),
        ("Created", 14), ("Days Open", 10), ("Age Bucket", 16),
    ]
    _header_row(ws, cols, row=4)

    def rating(score):
        if score >= 20: return "Critical"
        if score >= 15: return "High"
        if score >= 9:  return "Medium"
        return "Low"

    for i, r in enumerate(risks):
        days = (today - r.created_at.date()).days if r.created_at else 0
        inh  = r.likelihood * r.impact
        res  = (r.residual_likelihood or r.likelihood) * (r.residual_impact or r.impact)
        fill = _age_fill(days)
        row_data = [
            r.id, r.name,
            r.asset.name  if r.asset  else "",
            r.threat.name if r.threat else "",
            inh, res,
            rating(inh),
            (r.treatment or "").capitalize(),
            r.status.capitalize(),
            r.owner or "",
            r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
            days,
            _age_bucket(days),
        ]
        _body_row(ws, row_data, row=5 + i, alt=bool(i % 2), fill_override=fill)
        # Bold white text for 365+ rows
        if days > 365:
            for col in range(1, len(cols) + 1):
                ws.cell(row=5 + i, column=col).font = Font(name="Calibri", size=10, color="FFFFFF")

    # Sheet 2 — Aging Summary
    ws2 = wb.create_sheet("Aging Summary")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    _title_block(ws2, "Risk Aging Summary", f"As-of {today}")

    cols2 = [("Age Bucket", 18), ("Risk Count", 12), ("% of Total", 14)]
    _header_row(ws2, cols2, row=3)

    bucket_counts = {b: 0 for b in AGE_BUCKET_LABELS}
    for r in risks:
        days = (today - r.created_at.date()).days if r.created_at else 0
        bucket_counts[_age_bucket(days)] += 1

    for j, label in enumerate(AGE_BUCKET_LABELS):
        count = bucket_counts[label]
        pct   = f"{round(count / len(risks) * 100)}%" if risks else "0%"
        fill  = AGE_FILLS[j]
        _body_row(ws2, [label, count, pct], row=4 + j, alt=False, fill_override=fill)
        if j == 5 and count > 0:  # 365+ dark red — white font
            for col in range(1, 4):
                ws2.cell(row=4 + j, column=col).font = Font(name="Calibri", size=10, color="FFFFFF")

    AuditService(db).log("EXPORT_GENERATED", actor=current_user, resource_type="Export",
                         resource_name="risk_aging", request=request)
    return _xlsx_response(wb, f"risk_aging_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")
