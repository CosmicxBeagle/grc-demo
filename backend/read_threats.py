"""
Read threat names (row 1 header) and cell comments from columns BR–DB
in the SCF Excel file. Print everything so we can see what's there.
"""
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "scf_import.xlsx"

# Column letters → 0-based index helper
def col_letter_to_idx(letters):
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1   # 0-based

COL_START = col_letter_to_idx("BR")   # 69
COL_END   = col_letter_to_idx("DB")   # 105
print(f"Reading columns BR ({COL_START}) to DB ({COL_END}) — {COL_END - COL_START + 1} columns")

# Load WITHOUT read_only so comments are accessible
print(f"Loading {XLSX} (this may take a moment)...")
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["3 Frameworks Combined"]
print(f"Sheet loaded. Total cols: {ws.max_column}, Total rows: {ws.max_row}")

# Row 1 = headers (1-based in openpyxl)
for col_idx in range(COL_START + 1, COL_END + 2):   # openpyxl is 1-based
    cell = ws.cell(row=1, column=col_idx)
    col_letter = cell.column_letter
    value   = str(cell.value).strip() if cell.value else "(empty)"
    comment = cell.comment.text.strip() if cell.comment else "(no comment)"
    print(f"\n{'='*60}")
    print(f"Col {col_letter}  |  {value}")
    print(f"Comment:\n{comment}")

wb.close()
print("\nDone.")
